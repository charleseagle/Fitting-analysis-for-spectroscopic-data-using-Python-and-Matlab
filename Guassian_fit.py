from __future__ import division
import numpy as np
from numpy import *
import matplotlib.pyplot as plt
import scipy
import openpyxl as px
from openpyxl import Workbook
from lmfit import minimize, Parameters, report_fit


def gauss_dataset(x, params):
    #A, mu, sigma = p
    A = params['A'].value
    mu = params['mu'].value
    sigma = params['sigma'].value
    c = params['c'].value
    return A*np.exp(-(x-mu)**2/(2.*sigma**2))+c

#def gauss2(x, *args):
#    m1, m2, m3, s1, s2, s3, k1, k2, k3 = args
#    ret = k1*scipy.stats.norm.pdf(x, loc=m1 ,scale=s1)
#    ret += k2*scipy.stats.norm.pdf(x, loc=m2 ,scale=s2)
#    ret += k3*scipy.stats.norm.pdf(x, loc=m3 ,scale=s3)
#    return ret
#
#def gauss3(x, a, b, c, d, x0):
#    return a*np.exp(-(x-x0)**2/(2*d**2)) + c





W = px.load_workbook('E:\Python\Practice\UVvis_spec.xlsx', use_iterators = True)
p = W.get_sheet_by_name(name = 'Sheet1')


a=[]

for row in p.iter_rows():
    for k in row:
        a.append(k.internal_value)

b = len(row)
c = int(len(a)/b)
 
aa = np.resize(a, [c,b])


minwave = input('Input mininum wavelength for fit range:');
maxwave = input('Input maximum wavelength for fit range:');
#basewave = input ('Input wavelength for baseline subtraction (for quick Gaussian fit):');
mini = 0;
maxi = 0;
basei = 0;



'''Find the indices of minwave, maxwave, and basewave'''

for i in range(1,c,1):
    if aa[i,0] < minwave or mini == 0:
        mini = i
    if aa[i,0] < maxwave or maxi == 0:
        maxi = i
        



def objective(params, x, data):
    """ calculate total residual for fits to several data sets held
    in a 2-D array, and modeled by Gaussian functions"""
    ndata, nx = data.shape
    resid = 0.0*data[:]
    # make residual per data set
    resid[:, 1] = data[:, 1] - gauss_dataset(x, params)
    # now flatten this to a 1D array, as minimize() needs
    return resid.flatten()

d = aa.transpose()
cc = d[:,0]
dd = np.zeros((b,2))
# Equations and initial Parameterss.
for i in range(1,b):
    
    max_intensity = 0
    for j in range(mini,maxi,1):
        if aa[j,i] > max_intensity:
            max_intensity = aa[j,i]
            mini_fit = j

    
    xdata = aa[mini_fit:maxi,0]
    ydata = aa[mini_fit:maxi,i]
    
    fit_params = Parameters()
    fit_params.add('A', value = max_intensity)
    fit_params.add('mu', value = aa[mini_fit,0], vary = False)
    fit_params.add('sigma', value = 60)
    fit_params.add('c', value = 0.1)
    x = xdata
    y = ydata
    
   
    bb= column_stack((x,y))
    
    minimize(objective, fit_params, args=(x, bb))
    report_fit(fit_params)
    dd[i,0] = fit_params['mu'].value 
    dd[i,1] = 2.35482*fit_params['sigma'].value
    
    # plot the data sets and fits
    plt.figure()
    y_fit = gauss_dataset(x, fit_params)
    plt.plot(x, y, 'o', x, y_fit-0.08, '-')
    plt.show()

'''Generate an array including temperature (first row), peak position(second row)
and FWHM(third row)'''

ee = column_stack((cc,dd)) 
ee = np.delete(ee, (0), axis=0)

'''Plot plasmon wavelength, plasmon absorbance, and FWHM.'''
    
'''Write data to xlsx files'''


my_list = ee.tolist()
    
wb=px.Workbook()
ws = wb.active

for col in my_list:
    ws.append(col)

wb.save(filename='data.xlsx')


'''Write data to csv file.'''

#import csv
#
#fl = open('Plasmon.csv', 'w')
#
#writer = csv.writer(fl)
#writer.writerow(['Temperature', 'Peak position', 'FWHM']) #if needed
#for values in ee:
#    writer.writerow(values)
#fl.close()    

plt.figure()
plt.plot(ee[:,0], ee[:,1], 'o')
plt.savefig('temp.png')
plt.show()

plt.figure()
plt.plot(ee[:,0], ee[:,2], 's')

plt.savefig('temp1.jpg')
plt.show()













